from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
SITE="https://venezuela-photography-expedition-jwang815s-projects.vercel.app"
wb=Workbook(); ws=wb.active; ws.title="Guides"
INK="1B2430"; HEAD="10384A"; SOFT="5B6675"
thin=Side(style="thin",color="D9DEE5"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
def waurl(num): return "https://wa.me/"+num.replace("+","").replace(" ","")

ws["A1"]="Venezuela — Guide / Operator Outreach List"; ws["A1"].font=Font(name="Arial",size=15,bold=True,color=HEAD); ws.merge_cells("A1:M1")
ws["A2"]="Top block = researched & sourced June 2026 (confirm numbers before relying). Lower block = the 20 you supplied (2 exact duplicates removed). Blue 'Chat' opens WhatsApp. Right-hand columns are yours to track. Trip: 23 Jun–10 Jul 2026, 2 photographers."
ws["A2"].font=Font(name="Arial",size=9,italic=True,color=SOFT); ws.merge_cells("A2:M2"); ws["A2"].alignment=Alignment(wrap_text=True); ws.row_dimensions[2].height=30
headers=["#","Operator","Covers / base & specialty","WhatsApp","Open chat","Phone","Email","Website","Confidence","Notes","Contacted (date)","Quote (USD)","Status / next step"]
ws.append(headers); hr=3
for c in range(1,len(headers)+1):
    cell=ws.cell(row=hr,column=c); cell.font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor=HEAD); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
ws.row_dimensions[hr].height=30
conf_fill={"High":("E3F5EC","0C6B48"),"Med":("FBEEDE","8A5A12"),"Low":("FBE3DA","9A3217"),"Your list":("E7EEF6","2C4A6E")}

def write_row(r, idx, name, covers, wa, walink, phone, email, web, conf, notes):
    ws.cell(row=r,column=1,value=idx)
    ws.cell(row=r,column=2,value=name).font=Font(name="Arial",size=10,bold=True,color=INK)
    ws.cell(row=r,column=3,value=covers)
    ws.cell(row=r,column=4,value=wa)
    cc=ws.cell(row=r,column=5)
    if walink: cc.value="Chat ▶"; cc.hyperlink=(walink if walink.startswith("http") else waurl(walink)); cc.font=Font(name="Arial",size=10,color="0563C1",underline="single")
    else: cc.value="—"
    cc.alignment=Alignment(horizontal="center")
    ws.cell(row=r,column=6,value=phone)
    ec=ws.cell(row=r,column=7,value=email)
    if email and "@" in email and " " not in email.strip(): ec.hyperlink="mailto:"+email; ec.font=Font(name="Arial",size=10,color="0563C1",underline="single")
    wc=ws.cell(row=r,column=8,value=web)
    if web: wc.hyperlink=("https://"+web if not web.startswith("http") else web); wc.font=Font(name="Arial",size=10,color="0563C1",underline="single")
    fc,tc=conf_fill[conf]; cf=ws.cell(row=r,column=9,value=conf); cf.fill=PatternFill("solid",fgColor=fc); cf.font=Font(name="Arial",size=10,bold=True,color=tc); cf.alignment=Alignment(horizontal="center")
    ws.cell(row=r,column=10,value=notes)
    for c in range(1,14):
        cell=ws.cell(row=r,column=c); cell.border=border; cell.alignment=Alignment(vertical="top",wrap_text=(c in (2,3,10,13)),horizontal=cell.alignment.horizontal)
        if cell.font.name!="Arial": cell.font=Font(name="Arial",size=10,color=INK)
    ws.row_dimensions[r].height=42

R=[
("Osprey Expeditions","Nationwide / custom (all 5 regions)","+58 414 310 4491","584143104491","—","hola@ospreyexpeditions.com","ospreyexpeditions.com","High","Best all-rounder; English; active 2026; ideal for one full custom photo-trip quote."),
("MegaVenezuela","Nationwide / custom (all 5)","+58 424 216 1303","584242161303","—","via site form","megavenezuela.com","High","One-stop curator/DMC; English site; coordinates lodges across every leg."),
("Angel-Eco Tours","Canaima, Delta, custom (US office)","+58 424 156 6162","584241566162","US +1 212 656 1240","info@angel-ecotours.com","angel-ecotours.com","High","Higher-end custom; US booking office. NB VE line +58 414 310 4491 is shared with Osprey."),
("Araguato Expeditions","Delta, Canaima, Gran Sabana, Llanos, Roraima","not confirmed","","—","info@araguato.org","araguato.org","Med","Birding/wildlife multi-region operator; email first (no public WhatsApp found)."),
("Akanan (aadventures.net)","Nationwide / custom; US office","not confirmed","","US +1 678 862 9649","info@aadventures.net","aadventures.net","Med","US-fronted VE operator; confirm 2026 VE departures. Ignore old akanan.com."),
("Cacao Travel Group","Nationwide DMC (own lodges E/S)","not confirmed","","—","info@transcielos.com (B2B rep)","cacaotravel-venezuela.com","Low","Established DMC but site intermittently down; reach via TransCielos (Germany)."),
("Waku Lodge","Canaima / Angel Falls","+58 414 868 6659","584148686659","—","via site form","wakulodge.com","High","Premium Canaima lagoon lodge; photography-friendly base; Angel Falls + overflight."),
("Backpacker-Tours","Canaima, Roraima, Gran Sabana, Delta, Roques","+58 414 886 7227","584148867227","+58 289 995 1524","info@backpacker-tours.com","backpacker-tours.com","Med","Santa Elena trekking specialists; DE/EN/ES/PT; mid-budget."),
("Orinoco Delta Lodge","Orinoco Delta","+58 412 855 6225","584128556225","—","reservaciones@orinocodeltalodge.com","orinocodeltalodge.com","High","All-inclusive Warao-guided delta lodge; strong reviews; English+Spanish."),
("Campamento Mis Palafitos","Orinoco Delta","+58 424 282 8820","584242828820","+58 212 284 2015","via agencies","orinoco.travel/mis-palafitos-14","Med","Stilt-village delta camp; Spanish-first; boat + Warao excursions."),
("Orinoco Eco Camp","Orinoco Delta (+ Delta+Angel combos)","+58 414 091 4844 (likely WA)","584140914844","—","via site form","orinoco-eco-camp.com","Med","Eco/small-group; multilingual site. Try the mobile on WhatsApp; else the form."),
("Guamanchi Expeditions","Mérida / Andes (+ add-ons)","+58 424 775 2732","584247752732","+58 274 252 2080","info@guamanchi.com","guamanchi.com","High","Mérida trekking + own posadas (incl. by the Mukumbarí). English+Spanish."),
("Arassari Trek","Mérida / Andes","+58 414 746 3569 (likely WA)","584147463569","+58 274 252 5879","info@arassari.com","arassari.com","High","Well-known Mérida adventure operator (Pico Bolívar, rafting, canyoning)."),
("Natoura Travel & Adventure","Mérida + Delta + nationwide","click-to-chat (number hidden)","https://wa.me/message/AO2PD6QCFWWRO1","+58 274 252 4216","info@natoura.com","natoura.com","Med","Since 1991; bilingual; can quote Andes + Delta in one thread."),
("Programa Andes Tropicales","Mérida / Andes community posadas","not confirmed","","—","mucuposadas@gmail.com","andestropicales.org","Med","NGO community lodges; authentic páramo/culture; arrange ahead (Spanish)."),
("Viajando a Los Roques","Los Roques + multi-region","+58 424 411 6801","584244116801","—","Reservas@ValenciaLosRoques.com","viajandoalosroques.com","High","20+ yrs; all-inclusive Roques packages + other legs; online quoter."),
("WAKU Tours","Los Roques","+58 414 579 6887","584145796887","—","via site form","wakutours.com","High","Packaged Los Roques trips (Posada Acquamarina etc.); Caracas-based."),
("Posada Acquamarina","Los Roques (on-island stay)","not confirmed","","—","giorgio.serloni@gmail.com","posadaacquamarina.com","Low","Owner-run Gran Roque posada; DM FB/IG @acquamarinalosroques or book via WAKU."),
]
r=hr+1
for i,row in enumerate(R,1): write_row(r,i,row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8]); r+=1

# divider
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=13)
dc=ws.cell(row=r,column=1,value="Added from your list — 20 supplied, 2 exact duplicates removed (megaVENEZUELA, Osprey). Numbers as you provided; not independently verified.")
dc.font=Font(name="Arial",size=9,bold=True,italic=True,color="2C4A6E"); dc.fill=PatternFill("solid",fgColor="E7EEF6"); dc.alignment=Alignment(vertical="center",wrap_text=True)
ws.row_dimensions[r].height=24; r+=1

U=[
("Carmen Rondón (Paseo Macuto)","La Guaira historic centre, Fortines","+58 414 571 3310","584145713310",""),
("Transportes a Galipán (@galipanen4ruedas)","Galipán / Ávila 4x4 full-days","+58 412 972 3068","584129723068",""),
("Francisco Zambrano (@pachoencaracas)","Galipán ATV, Mirador Caracas, Río San José","+58 412 276 5938","584122765938",""),
("Viajeros Venezuela","Barquisimeto, Lara — city tours","+58 412 673 8548","584126738548",""),
("Valencia en Tranvía","Valencia heritage bus circuit","+58 414 404 7979","584144047979",""),
("Experiencia Tours","Caracas (Plaza Venezuela/El Marqués) — nationwide tickets + full-days","+58 414 206 9502","584142069502",""),
("CruzdeBolívar","Ciudad Bolívar — gateway to Canaima","+58 285 651 9256","","Landline — call; may not be on WhatsApp."),
("Turismo Venezuela","Cumaná, Sucre — coastal tours","+58 416 582 5369","584165825369",""),
("Trípode tours","Miranda state day trips","+58 424 189 2657","584241892657",""),
("Campamento Parakaupa (Sapo)","Canaima NP — Laguna sector","+58 416 931 9140","584169319140",""),
("The Rama Yacht","Marina Bahía Redonda, Anzoátegui — boat tours","+58 424 887 0144","584248870144",""),
("Hike Venezuela — Thorsten 'Tino' Blum","Andes, Roraima, Canaima treks","+58 412 359 1797","584123591797","English/German-speaking trekking guide."),
("Hike Venezuela — Andreas","Same routes (Andes / Roraima / Canaima)","+58 412 795 8147","584127958147","German/English-speaking."),
("HectanaTravels","Choroní, coastal pueblos","+58 412 887 2334","584128872334",""),
("Aladins Aventura","Morrocoy / La Ciénaga full-days","+58 426 814 6659","584268146659",""),
("VenEventura Tours","Mérida — Andes, Teleférico Mukumbarí","+58 414 606 0792","584146060792",""),
("Mochilero Extremo","Canaima / Salto Ángel 5D4N","+58 412 500 5598 / +58 416 300 5983","584125005598",""),
("Pinero Tours","Mérida, Los Roques — flights & tours","+58 414 415 7372","584144157372",""),
]
for j,row in enumerate(U,1): write_row(r,len(R)+j,row[0],row[1],row[2],row[3],"","","","Your list",row[4]); r+=1

widths=[4,27,30,21,11,16,28,24,11,40,15,13,24]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A4"

ms=wb.create_sheet("Outreach message")
ms["A1"]="Ready-to-send outreach (paste into WhatsApp / email)"; ms["A1"].font=Font(name="Arial",size=13,bold=True,color=HEAD)
ms["A3"]="Your itinerary link (send this):"; ms["A3"].font=Font(name="Arial",size=10,bold=True)
ms["A4"]=SITE; ms["A4"].hyperlink=SITE; ms["A4"].font=Font(name="Arial",size=10,color="0563C1",underline="single")
EN=("Hello! I'm planning an 18-day photography trip in Venezuela, 23 June–10 July 2026, for 2 photographers: "
"Caracas → Orinoco Delta → Mérida → Canaima/Angel Falls → Los Roques. The full day-by-day plan is here:\n"+SITE+"\n\n"
"Could you (1) send a quote for the legs you cover, and (2) suggest any adjustments — flight timing, feasibility, lodges? "
"Two travelers, paying in USD, flexible on accommodation. Thank you!")
ES=("¡Hola! Estoy planificando un viaje de fotografía de 18 días en Venezuela, del 23 de junio al 10 de julio de 2026, para 2 fotógrafos: "
"Caracas → Delta del Orinoco → Mérida → Canaima/Salto Ángel → Los Roques. El itinerario completo, día por día, está aquí:\n"+SITE+"\n\n"
"¿Podrían (1) enviarme una cotización por los tramos que cubren y (2) sugerir ajustes — horarios de vuelos, viabilidad, posadas? "
"Somos 2 viajeros, pago en USD, flexibles con el alojamiento. ¡Muchas gracias!")
ms["A6"]="English"; ms["A6"].font=Font(name="Arial",size=11,bold=True)
ms["A7"]=EN; ms["A7"].alignment=Alignment(wrap_text=True,vertical="top"); ms.merge_cells("A7:H14")
ms["A16"]="Español"; ms["A16"].font=Font(name="Arial",size=11,bold=True)
ms["A17"]=ES; ms["A17"].alignment=Alignment(wrap_text=True,vertical="top"); ms.merge_cells("A17:H24")
ms.column_dimensions["A"].width=22
for col in "BCDEFGH": ms.column_dimensions[col].width=14
for rr in (7,17): ms.row_dimensions[rr].height=120

out="/sessions/confident-dreamy-cori/mnt/outputs/Venezuela_Guide_Outreach_List.xlsx"
wb.save(out); print("saved · total operators:",len(R)+len(U))
